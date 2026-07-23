import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

import arabic_reshaper
from bidi.algorithm import get_display

from sqlalchemy.ext.asyncio import AsyncSession
from bot.services.report_service import ReportService
from bot.services.finance_service import FinanceService
from bot.services.student_service import StudentService
from bot.services.sponsor_service import SponsorService


import os
import urllib.request
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

FONT_PATH = os.path.join(os.path.dirname(__file__), "..", "assets", "fonts", "Amiri-Regular.ttf")


def register_arabic_font() -> str:
    """Registers an Arabic-compatible TrueType font for ReportLab."""
    font_name = "Amiri"
    font_dir = os.path.dirname(FONT_PATH)
    
    if not os.path.exists(font_dir):
        os.makedirs(font_dir, exist_ok=True)

    if not os.path.exists(FONT_PATH):
        # Fallback to Windows system font if available
        win_arial = "C:\\Windows\\Fonts\\arial.ttf"
        if os.path.exists(win_arial):
            pdfmetrics.registerFont(TTFont(font_name, win_arial))
            return font_name
        
        # Download Amiri-Regular font if missing
        try:
            url = "https://raw.githubusercontent.com/google/fonts/main/ofl/amiri/Amiri-Regular.ttf"
            urllib.request.urlretrieve(url, FONT_PATH)
        except Exception:
            pass

    if os.path.exists(FONT_PATH):
        pdfmetrics.registerFont(TTFont(font_name, FONT_PATH))
        return font_name

    return "Helvetica"


def reshape_ar(text: str) -> str:
    """Helper to reshape and reorder Arabic text for ReportLab PDF rendering."""
    if not text:
        return ""
    reshaped_text = arabic_reshaper.reshape(str(text))
    return get_display(reshaped_text)


class ExportService:
    @staticmethod
    async def generate_excel_report(session: AsyncSession) -> io.BytesIO:
        wb = Workbook()
        
        # Styles
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        title_font = Font(name="Segoe UI", size=14, bold=True, color="1F4E79")
        align_center = Alignment(horizontal="center", vertical="center")
        border_thin = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9")
        )

        # 1. Dashboard / Summary Sheet
        ws_sum = wb.active
        ws_sum.title = "الملخص المالي"
        ws_sum.views.sheetView[0].rightToLeft = True

        stats = await ReportService.get_dashboard_stats(session)
        
        ws_sum.append(["تقرير المالي - دفعة يمن سايبر"])
        ws_sum.merge_cells("A1:B1")
        ws_sum["A1"].font = title_font

        ws_sum.append([])
        summary_rows = [
            ("إجمالي الإيرادات", stats["total_income"]),
            ("إجمالي المصروفات", stats["total_expenses"]),
            ("الرصيد الحالي", stats["current_balance"]),
            ("إجمالي عدد الطلاب", stats["total_students"]),
            ("عدد الطلاب المسددين", stats["paid_students"]),
            ("عدد الطلاب غير المسددين", stats["unpaid_students"]),
            ("إجمالي المتبقي على الطلاب", stats["total_student_remaining"]),
            ("أكبر داعم", stats["top_sponsor"]),
            ("قيمة دعم أكبر داعم", stats["top_sponsor_amount"]),
        ]
        
        for label, val in summary_rows:
            ws_sum.append([label, val])
            
        for row in ws_sum.iter_rows(min_row=3, max_col=2):
            for cell in row:
                cell.border = border_thin
                cell.alignment = align_center

        # 2. Incomes Sheet
        ws_inc = wb.create_sheet(title="الإيرادات")
        ws_inc.views.sheetView[0].rightToLeft = True
        inc_headers = ["رقم العملية", "النوع", "المبلغ", "التاريخ", "المستلم", "ملاحظات"]
        ws_inc.append(inc_headers)
        for col_num, h_text in enumerate(inc_headers, 1):
            cell = ws_inc.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center

        incomes = await FinanceService.get_latest_incomes(session, limit=500)
        for inc in incomes:
            ws_inc.append([
                inc.op_number,
                inc.category.value,
                float(inc.amount),
                inc.date.strftime("%Y-%m-%d %H:%M"),
                inc.recipient_name,
                inc.notes or ""
            ])

        # 3. Expenses Sheet
        ws_exp = wb.create_sheet(title="المصروفات")
        ws_exp.views.sheetView[0].rightToLeft = True
        exp_headers = ["رقم العملية", "السبب", "المبلغ", "التاريخ", "المستفيد", "ملاحظات"]
        ws_exp.append(exp_headers)
        for col_num, h_text in enumerate(exp_headers, 1):
            cell = ws_exp.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center

        expenses = await FinanceService.get_latest_expenses(session, limit=500)
        for exp in expenses:
            ws_exp.append([
                exp.op_number,
                exp.reason,
                float(exp.amount),
                exp.date.strftime("%Y-%m-%d %H:%M"),
                exp.beneficiary_name,
                exp.notes or ""
            ])

        # 4. Students Sheet
        ws_stu = wb.create_sheet(title="الطلاب")
        ws_stu.views.sheetView[0].rightToLeft = True
        stu_headers = ["الاسم", "الرقم الأكاديمي", "الهاتف", "المبلغ المطلوب", "المدفوع", "المتبقي", "حالة السداد"]
        ws_stu.append(stu_headers)
        for col_num, h_text in enumerate(stu_headers, 1):
            cell = ws_stu.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center

        students = await StudentService.get_all_students(session)
        for st in students:
            ws_stu.append([
                st.name,
                st.student_code,
                st.phone or "",
                float(st.total_required),
                float(st.total_paid),
                float(st.remaining_amount),
                st.status.value
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    async def generate_pdf_report(session: AsyncSession) -> io.BytesIO:
        font_name = register_arabic_font()

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        elements = []

        styles = getSampleStyleSheet()
        normal_style = styles["Normal"]
        title_style = ParagraphStyle(
            "ArabicTitle",
            parent=normal_style,
            fontName=font_name,
            fontSize=18,
            leading=22,
            alignment=1, # Center
            textColor=colors.HexColor("#1F4E79")
        )

        subtitle_style = ParagraphStyle(
            "ArabicSubTitle",
            parent=normal_style,
            fontName=font_name,
            fontSize=12,
            leading=16,
            alignment=1,
            textColor=colors.HexColor("#555555")
        )

        cell_style = ParagraphStyle(
            "ArabicCell",
            parent=normal_style,
            fontName=font_name,
            fontSize=11,
            leading=14,
            alignment=2 # Right aligned
        )

        stats = await ReportService.get_dashboard_stats(session)

        # Header Title
        title_p = Paragraph(reshape_ar("التقرير المالي - دفعة تخرج يمن سايبر"), title_style)
        date_p = Paragraph(reshape_ar(f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d')}"), subtitle_style)
        elements.append(title_p)
        elements.append(date_p)
        elements.append(Spacer(1, 20))

        # Financial Summary Table
        table_data = [
            [Paragraph(reshape_ar("القيمة"), cell_style), Paragraph(reshape_ar("البيان"), cell_style)],
            [Paragraph(reshape_ar(f"{stats['total_income']:,.0f} ريال"), cell_style), Paragraph(reshape_ar("إجمالي الإيرادات"), cell_style)],
            [Paragraph(reshape_ar(f"{stats['total_expenses']:,.0f} ريال"), cell_style), Paragraph(reshape_ar("إجمالي المصروفات"), cell_style)],
            [Paragraph(reshape_ar(f"{stats['current_balance']:,.0f} ريال"), cell_style), Paragraph(reshape_ar("الرصيد الحالي"), cell_style)],
            [Paragraph(reshape_ar(f"{stats['total_students']} طالب"), cell_style), Paragraph(reshape_ar("عدد الطلاب"), cell_style)],
            [Paragraph(reshape_ar(f"{stats['paid_students']} طالب"), cell_style), Paragraph(reshape_ar("الطلاب المسددون"), cell_style)],
            [Paragraph(reshape_ar(f"{stats['unpaid_students']} طالب"), cell_style), Paragraph(reshape_ar("الطلاب المتبقون"), cell_style)],
            [Paragraph(reshape_ar(f"{stats['total_student_remaining']:,.0f} ريال"), cell_style), Paragraph(reshape_ar("المتبقي على الطلاب"), cell_style)],
        ]

        t = Table(table_data, colWidths=[200, 250])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (1,0), colors.HexColor("#1F4E79")),
            ('TEXTCOLOR', (0,0), (1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
            ('FONTNAME', (0,0), (-1,0), font_name),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('BACKGROUND', (0,1), (-1,-1), colors.HexColor("#F2F2F2")),
            ('GRID', (0,0), (-1,-1), 1, colors.HexColor("#D9D9D9")),
        ]))
        elements.append(t)

        doc.build(elements)
        output.seek(0)
        return output

